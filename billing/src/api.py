from src import app
from flask import request,send_file,Response
from enum import Enum
import json
import openpyxl
from pathlib import Path
import mysql.connector
import requests
import xlwt
import io


class LOG_TYPE(Enum):
    INFO=0
    ERROR=1
    WARNING=2

def xlsx_to_array(dir,filename):
    xlsx_file = Path(dir, filename)
    wb_obj = openpyxl.load_workbook(xlsx_file) 
    sheet = wb_obj.active
    return sheet.iter_rows(min_row=2)

type_desc = ["[INFO]","[ERROR]","[WARNING]"]


def LOG(str,TYPE):
    with open("log","a+") as f:
        f.write(f'{type_desc[TYPE.value]} {str}\n')


def connect():
    mydb=None
    LOG("Connecting to Database...",LOG_TYPE.INFO)
    
    try:
        mydb = mysql.connector.connect(
            host="billingDB",
            user="root",
            password="catty",
            database="billdb"
        )
    except:
        LOG("Connection failed!",LOG_TYPE.ERROR)
        return None

    LOG("Connection successful!",LOG_TYPE.INFO)
    return mydb

@app.route("/health")
def health():
    return "OK"

@app.route("/db_health")
def db_health():
    return ("BAD" if connect() == None else "OK")

@app.route('/provider',methods=['POST'])
def Insert_provider():

    if request.form['name'] == None or len(request.form['name']) == 0:
        LOG("Invalid data provided! Cannot insert into Provider table",LOG_TYPE.ERROR)
        return "BAD"
    
    #Open connection to the database
    db = connect()
    if db == None:
        return "BAD"

    LOG("Inserting into Provider table...",LOG_TYPE.INFO)

    try:
        mycursor = db.cursor()
        #Query SQL
        sql = 'INSERT INTO Provider(name) VALUES(%s)'
        val = (request.form['name'],)
        
        mycursor.execute(sql,val)
        
        db.commit()#Commit changes of the SQL Query

        LOG("Inserted new record into Provider table",LOG_TYPE.INFO)

        return json.dumps({'id':mycursor.lastrowid}), 201, {'Content-Type':'application/json'}
    except ValueError:
        LOG("Failed to insert data into Provider Table",LOG_TYPE.ERROR)
    finally:
        LOG("Closing connection to database...",LOG_TYPE.INFO)
        db.close()#Close connection

    return "BAD"


@app.route('/provider/<id>',methods=['PUT'])
def Update_provider(id):
    data = request.form['name']
    
    if len(request.form['name']) == 0:
        LOG("Invalid data provided! Cannot Update Provider table",LOG_TYPE.ERROR)
        return "BAD",400
    
    #Open connection to the database
    db = connect()
    if db == None:
        return "Database connection failed",500

    try:
        mycursor = db.cursor()
        
        sql = "SELECT * FROM Provider WHERE id=%s"
        val = (id,)

        mycursor.execute(sql,val)
        
        if len(mycursor.fetchall()) == 0:
            #TODO Add log here
            return "Provider not found",404


        sql = "UPDATE Provider SET name=%s WHERE id=%s"
        val = (request.form['name'],id)

        mycursor.execute(sql,val)
        db.commit()

        LOG("Updated Provider table successfully!",LOG_TYPE.INFO)

        return "OK"
    except:
        LOG("Failed to Update Provider table",LOG_TYPE.ERROR)
        pass
    finally:
        LOG("Closing connection to database...",LOG_TYPE.INFO)
        db.close()
    
    return "BAD"

@app.route("/rates",methods=['POST'])
def rates():
   data = request.form
   rates = xlsx_to_array("/in",data['filename'])
   db = connect()
   mycursor = db.cursor()
   for row in rates:
    sql = 'SELECT * FROM Rates WHERE product_id = %s AND scope = %s'
    val = (row[0].value,row[2].value)
    mycursor.execute(sql,val)
    myresult = mycursor.fetchall()
    if len(myresult) == 0:
        sql = 'insert into Rates (product_id,rate,scope) VALUES (%s,%s,%s)'
        val = (row[0].value,row[1].value,row[2].value)
    else:
        sql = 'UPDATE Rates SET rate = %(rate)s WHERE product_id = %(product_id)s AND scope = %(scope)s'
        val = {'product_id':row[0].value,
                'rate':row[1].value,
                'scope':row[2].value}
    mycursor.execute(sql, val)
    db.commit()
    db.close()

    return "OK",201

@app.route('/rates',methods=['GET'])
def Download_RatesXL():
    db = connect()
    if db == None:
        return "BAD"

    mycursor = db.cursor()
    mycursor.execute('SELECT * FROM Rates')
    myresult = mycursor.fetchall()

    #output in bytes
    output = io.BytesIO()
    #create WorkBook object
    workbook = xlwt.Workbook()
    #add a sheet
    sh = workbook.add_sheet('rates')

    #add headers
    sh.write(0, 0, 'product_id')
    sh.write(0, 1, 'rate')
    sh.write(0, 2, 'scope')

    idx = 0
    for row in myresult:
        sh.write(idx+1, 0, str(row[0]))
        sh.write(idx+1, 1, row[1])
        sh.write(idx+1, 2, str(row[2]))
        idx += 1

    workbook.save(output)
    output.seek(0) 

    return Response(output, mimetype="application/ms-excel", headers={"Content-Disposition":"attachment;filename=rates-new.xls"}) 

@app.route('/truck',methods=['POST'])
def Insert_truck():

    #Open connection to the database
    db = connect()
    if db == None:
        return "BAD"

    try:
        mycursor = db.cursor()
        sql = "INSERT INTO Trucks (id,provider_id) VALUES (%s,%s)"
        val = (request.form['id'],request.form['provider'])
        
        mycursor.execute(sql,val)
        db.commit()

        LOG("Inserted new record into Trucks table successfully!",LOG_TYPE.INFO)

        return "OK",201
    except:
        LOG("Failed to insert new record into Trucks table",LOG_TYPE.ERROR)
        pass
    finally:
        LOG("Closing connection to database...",LOG_TYPE.INFO)
        db.close()

    return "BAD"

@app.route("/truck/<id>",methods=['PUT'])
def Update_ProviderID(id):

    #Open connection to the database
    db = connect()
    if db == None:
        return "BAD"
    
    try:
        mycursor = db.cursor()
        
        sql = "UPDATE Trucks SET provider_id=%s WHERE id=%s"
        val = (request.form['provider_id'],id)

        mycursor.execute(sql,val)
        db.commit()

        LOG("Updated Trucks table Column[provider_id] successfully!",LOG_TYPE.INFO)

        return "OK"
    except:
        LOG("Failed to Update Trucks table",LOG_TYPE.ERROR)
        pass
    finally:
        LOG("Closing connection to database...",LOG_TYPE.INFO)
        db.close()
    
    return "OK"


#Mock-up routes
@app.route('/item',methods=['GET'])
def Get_Item():
    res=''
    data=None
    with open("json-mock.json") as f:
        data = json.load(f)
    return json.dumps(data),200,{'Content-Type': 'application/json'}

@app.route("/truck/<id>",methods=['GET'])
def Get_Truck(id):
    connection = connect()
    mycursor = connection.cursor()

    sql_query = "SELECT * FROM Trucks WHERE id=%s"
    val = (id,)
    mycursor.execute(sql_query,val)
    results = mycursor.fetchall()
    
    LOG(f'{results}',LOG_TYPE.INFO)

    #TODO Perform check to see if it's found
    if len(results) == 0: #Empty means we don't have a truck with this license plate
        return 404
    
    return requests.get('http://localhost:5000/item',{'t1':'','t2':''}).content

@app.route("/weight",methods=['GET'])
def Get_Weight():
    data=None
    with open("json-mock-weights.json") as f:
        data = json.load(f)
    return json.dumps(data),200,{'Content-Type': 'application/json'}

@app.route("/bill/<id>",methods=['GET'])
def Get_Bill(id):
    connection = connect()
    if connection == None:
        return "BAD"
    
    name = "ad"
    date_from = request.args.get("from")
    date_to = request.args.get("to")
    truckCount=0
    total=0

    LOG(f'Parameter 1: {date_from} Parameter 2: {date_to}',LOG_TYPE.INFO)

    mycursor = connection.cursor()
    
    sql_query = "SELECT name FROM Provider WHERE id=%s"
    val = (id,)
    mycursor.execute(sql_query,val)
    result_arr = mycursor.fetchall()
    if len(result_arr) == 0:
        return "Provider not found",404
    
    name = result_arr[0][0]

    sql_query = "SELECT id FROM Trucks WHERE provider_id=%s"
    val = (id,)
    mycursor.execute(sql_query,val)
    result_arr = mycursor.fetchall()
    if len(result_arr) == 0:
        return "Not trucks were registered to this provider",200
    
    truckCount = len(result_arr)

    #We expect to get a json array for all trucks that went out
    weighted_containers = json.loads(requests.get("http://localhost:5000/weight",{'from':date_from,'to':date_to,'filter':'out'}).content)

    #TODO Number of sessions for each product
    #TODO Get total weight of each product
    providers_sessions = []

    #Get all sessions that are specific to this provider
    for truck in weighted_containers:
        for truck_id in result_arr[0]:
            LOG(truck['id'],LOG_TYPE.INFO)
            if truck['id'] == truck_id:
                providers_sessions.append(truck)

    providers_products=[]
    
    for truck in providers_sessions:
        flag=0
        for product in providers_products:
            if truck['produce'] == product['name']:
                product['count'] += 1
                product['amount'] += truck['neto']
                product['pay'] += product['rate']*truck['neto']
                total += product['rate']*truck['neto']
                flag=1
        if flag == 0:#No Products were found in the list we add it
            #Fetch the rates for this specific provider
            sql_query = "SELECT rate FROM Rates WHERE product_id=%s AND scope=%s"
            val = (truck['produce'],id)
            mycursor.execute(sql_query,val)
            rates_arr = mycursor.fetchall()
            if len(rates_arr) == 0:#We didn't find speific rate for this provider we have to look for ALL scope
                sql_query = "SELECT rate FROM Rates WHERE product_id=%s AND scope='ALL'"
                val = (truck['produce'],)
                mycursor.execute(sql_query,val)
                rates_arr = mycursor.fetchall()
            #TODO for now we assume product is found in rates but that's not always the case so we have to add another check

            providers_products.append({
                "name": truck['produce'],
                "count" : 1,
                "amount" : truck['neto'],
                "rate": int(rates_arr[0][0]),
                "pay" : truck['neto']*int(rates_arr[0][0])
            })
            total += int(rates_arr[0][0])*truck['neto']
    
    
    connection.close()
    recipet = {
        "id": id,
        "name": name,
        "from": date_from,
        "to": date_to,
        "truckCount": truckCount,
        "sessionCount": len(providers_sessions),
        "products": providers_products,
        "total": total
    }
    return json.dumps(recipet),200,{'Content-Type':'application/json'}
